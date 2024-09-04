import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import json
import openpyxl
import os
import subprocess
import re
import pandas as pd
import sys
import time
from googleapiclient.http import MediaIoBaseDownload
import io


# File paths
excel_file_path = 'EmployeeListSchedule.xlsx'
json_file_path = 'json/formatted_users.json'


def print_colored(text, color):
    colors = {
        'red': '\033[91m',
        'white': '\033[92m',
        'yellow': '\033[93m',
        'blue': '\033[94m',
        'magenta': '\033[95m',
        'white': '\033[96m',
        'white': '\033[97m',
        'reset': '\033[0m'
    }
    print(colors.get(color, colors['reset']) + text + colors['reset'])


def prompt_user():
    response = input("Do you want to update the employee list? (y/n): ")
    return response.strip().lower() == 'y'


def update_employee_list():
    # Read the Excel file
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file_path)
    except FileNotFoundError:
        print(
            f"Error: The Excel file '{excel_file_path}' was not found...\nContinuing to update schedule.", "red")
        time.sleep(3)
        return

    # Load existing JSON data
    if os.path.exists(json_file_path):
        with open(json_file_path, 'r') as json_file:
            formatted_users = json.load(json_file)
    else:
        formatted_users = []

    # Create a dictionary for existing users using first and last names as keys
    existing_users = {}
    for user in formatted_users:
        first_name = user.get('firstName')
        last_name = user.get('lastName')
        if first_name and last_name:
            existing_users[(first_name, last_name)] = user

    # Office mapping
    office_mapping = {
        "Grafton": "Milwaukee",
        "Baraboo": "Madison",
        "Florida": "Rockford",
        "Stevens Point": "Fox Valley"
    }

    # Process each row in the DataFrame
    for index, row in df.iterrows():
        first_name = row['FirstName'].capitalize()
        last_name = row['LastName'].capitalize()
        # Convert employee number to string
        emp_num = str(row['EmployeeNumber'])
        office = row['OfficeName']

        # Apply office mapping
        office = office_mapping.get(office, office)

        # Create the username, display name, and determine office location
        username = first_name[0].lower() + last_name.lower()
        display_name = f"{first_name} {last_name[0]}"

        # Create a dictionary for the formatted user data
        user_data = {
            "username": username,
            "password": emp_num,
            "type": "user",
            "displayName": display_name,
            "office": office,
            "firstName": first_name,
            "lastName": last_name
        }

        # Update existing user or add new user based on first and last names
        existing_users[(first_name, last_name)] = user_data

    # Convert the dictionary back to a list
    updated_users = list(existing_users.values())

    # Write the updated user data to the JSON file
    with open(json_file_path, 'w') as json_file:
        json.dump(updated_users, json_file, indent=2)

    print(
        "Formatted user data has been saved to formatted_users.json")


# Main script execution
if prompt_user():
    update_employee_list()
else:
    print("Skipping employee list update...\n\n")


json_file_path = 'json/store_runs.json'

# Check if the file exists and then delete it
if os.path.exists(json_file_path):
    os.remove(json_file_path)

    # Define the current directory
directory = '.'

# List all files in the directory
files_in_directory = os.listdir(directory)

# Filter out Excel files
excel_files = [file for file in files_in_directory if file.endswith(
    '.xlsx') or file.endswith('.xls')]

# Delete each Excel file
for file in excel_files:
    os.remove(os.path.join(directory, file))


class StoreRun:
    def __init__(self, date, start_time, employee_list=None):
        self.date = date
        self.meet_time = []
        self.start_time = start_time
        self.inv_type = []
        self.store_name = []
        self.address = []
        self.link = []
        self.store_note = []
        self.employee_list = {}

    def add_employee(self, name, number, note, office):
        self.employee_list[name] = [number, note, office]

    def add_inv_type(self, inv_type):
        self.inv_type.append(inv_type)

    def add_store_name(self, store_name):
        self.store_name.append(store_name)

    def add_address(self, address):
        self.address.append(address)

    def add_link(self, link):
        self.link.append(link)

    def add_store_note(self, store_note, concat=False):
        if concat and self.store_note:
            self.store_note[-1] += f""" <br><hr id="note_separator">{store_note}"""
        else:
            self.store_note.append(store_note)

    def __str__(self):
        return f"date={self.date}, meet_time={self.meet_time}, start_time={self.start_time}, inv_type={self.inv_type}, store_name={self.store_name}, address={self.address}, link={self.link}, store_note={self.store_note}, employee_list={self.employee_list})"

    def to_dict(self):
        return {
            "date": self.date,
            "meet_time": self.meet_time,
            "start_time": self.start_time,
            "inv_type": self.inv_type,
            "store_name": self.store_name,
            "address": self.address,
            "link": self.link,
            "store_note": self.store_note,
            "employee_list": self.employee_list,
        }


def save_store_runs_to_json():
    # after creating and populating the store_run object
    json_file_path = 'json/store_runs.json'

    try:
        with open(json_file_path, 'r+') as file:
            # First we load existing data into a dict.
            file_data = json.load(file)
            # Join new_data with file_data inside emp_details
            file_data.append(store_run.to_dict())
            # Sets file's current position at offset.
            file.seek(0)
            # convert back to json.
            json.dump(file_data, file, indent=4)
    except FileNotFoundError:
        with open(json_file_path, 'w') as file:
            # If the file doesn't exist, create a new one and write the data.
            json.dump([store_run.to_dict()], file, indent=4)


script_dir = os.path.dirname(__file__)
config_path = os.path.join(script_dir, 'json/config.json')

# Load credentials from a JSON file
with open(config_path) as config_file:
    credentials = json.load(config_file)

# Set up the credentials for Google Sheets and Drive API
scope = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/documents'

]
creds = Credentials.from_service_account_info(credentials, scopes=scope)

# Initialize the clients for Sheets and Drive API
client = gspread.authorize(creds)
drive_service = build('drive', 'v3', credentials=creds)

# The name of the folder you want to search for
folder_name = 'Fake Milwaukee'

# Search for the folder by name to get its ID
query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}'"
folder_response = drive_service.files().list(q=query).execute()
folders = folder_response.get('files', [])

if folders:
    folder_id = folders[0]['id']  # Taking the first folder found

    # List all sheets in the specified folder using the folder ID
    doc_query = f"mimeType='application/vnd.google-apps.document' and '{folder_id}' in parents"
    doc_file_response = drive_service.files().list(q=doc_query).execute()
    doc_files = doc_file_response.get('files', [])

    for file in doc_files:
        file_id = file['id']
        file_name = file['name']

        # Request to export the Google Docs file as a .docx file
        request = drive_service.files().export_media(fileId=file_id,
                                                     mimeType='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        file_path = os.path.join(os.getcwd(), f"{file_name}.docx")

        # Download the file to the current directory
        with io.FileIO(file_path, 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                print(
                    f"Downloading {file_name}: {int(status.progress() * 100)}%")

    # List all sheets in the specified folder using the folder ID
    query = f"mimeType='application/vnd.google-apps.spreadsheet' and '{folder_id}' in parents"
    file_response = drive_service.files().list(q=query).execute()
    files = file_response.get('files', [])
    files = sorted(files, key=lambda x: x['name'])
    process_columns = [2, 6, 10, 14, 18, 22, 26]

    for file in files:
        # Open the spreadsheet by ID using gspread
        gsheet = client.open_by_key(file['id'])
        # Print the name of the spreadsheet
        print(f"Processing sheet: {gsheet.title}")

        # Select the worksheet by title
        worksheet = gsheet.sheet1

        # Extract data from Google Sheets
        data = worksheet.get_all_values()

        # Create a new Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Insert data into the Excel sheet maintaining formatting
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_data in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index).value = cell_data

        workbook.save(gsheet.title + '.xlsx')

        if workbook:
            for process_col in process_columns:

                # Starting from row 8 in column B
                min_row = 8
                current_state = 'searching'  # Initial state before setting any values
                break_outer_loop = False  # Flag to control breaking out of the outer loop
                header_value = sheet.cell(row=1, column=process_col - 1).value
                courtesy_meet = False

                # Iterate through the rows starting from min_row and process_col
                for row in sheet.iter_rows(min_row=min_row, max_row=500, min_col=process_col, max_col=process_col):
                    if break_outer_loop:
                        break
                    for cell in row:
                        value = cell.value
                        if value:
                            value = value.strip()
                        # Get the value of the cell to the left (number value)
                        number_value = sheet.cell(
                            row=cell.row, column=cell.column - 1).value
                        # Get the value of the cell to the right (note value)
                        note_value = sheet.cell(
                            row=cell.row, column=cell.column + 1).value
                        # Get the value of the cell two rows down (next store)
                        next_store = sheet.cell(
                            row=cell.row + 1, column=cell.column).value
                        # Handle empty days (weekends)
                        if cell.row == 8 and cell.value is None:
                            next_cell = sheet.cell(
                                row=cell.row + 1, column=cell.column).value
                            if next_cell is None:
                                for i in range(1, 20):  # Check the next 3 cells
                                    next_cell = sheet.cell(
                                        row=cell.row + i, column=cell.column).value
                                if next_cell is None:
                                    store_run = StoreRun(
                                        date=None, start_time=None)
                                    store_run.date = header_value
                                    save_store_runs_to_json()
                                    break_outer_loop = True
                                    break
                        elif cell.row == 8 and cell.value == '':
                            next_cell = sheet.cell(
                                row=cell.row + 1, column=cell.column).value
                            if next_cell == '':
                                for i in range(1, 20):  # Check the next 3 cells
                                    next_cell = sheet.cell(
                                        row=cell.row + i, column=cell.column).value
                                if next_cell == '':
                                    store_run = StoreRun(
                                        date=None, start_time=None)
                                    store_run.date = header_value
                                    save_store_runs_to_json()
                                    break_outer_loop = True
                                    break

                        # Assuming value is the string containing the meet times
                        if value and current_state == 'searching' and 'courtesy meet' in value.lower():
                            current_state == 'searching'
                            courtesy_meet = True
                        elif value and current_state == 'searching' and 'meet' in value.lower():
                            store_run = StoreRun(
                                date=None, start_time=None)
                            # Split the value by lines
                            lines = value.split('\n')
                            # Initialize a dictionary to hold meet times
                            meet_times_dict = {'M:': None,
                                               'IL:': None, 'FV:': None, 'MD:': None}
                            default_value = None
                            # Iterate over each line to find meet times
                            for line in lines:
                                if line.startswith('M:'):
                                    meet_times_dict['M:'] = line
                                elif line.startswith('IL:'):
                                    meet_times_dict['IL:'] = line
                                elif line.startswith('FV:'):
                                    meet_times_dict['FV:'] = line
                                elif line.startswith('MD:'):
                                    meet_times_dict['MD:'] = line
                                else:
                                    default_value = line.strip()
                            # Add meet times to the list in the specified order
                            store_run.meet_time = [
                                meet_times_dict['M:'] or default_value,
                                meet_times_dict['IL:'],
                                meet_times_dict['FV:'],
                                meet_times_dict['MD:']
                            ]

                            if courtesy_meet:
                                # Remove any None values from the list
                                store_run.meet_time = [
                                    time + '\n(COURTESY MEET)' for time in store_run.meet_time if time is not None]
                            else:
                                store_run.meet_time = [
                                    time for time in store_run.meet_time if time is not None]
                            current_state = 'found_meet'
                        elif value and current_state == 'searching' and 'leave' in value.lower():
                            if courtesy_meet:
                                store_run = StoreRun(
                                    date=None,  start_time=None)
                                store_run.meet_time = value + \
                                    '\n(COURTESY MEET)'
                            else:
                                store_run = StoreRun(
                                    date=None,  start_time=None)
                                store_run.meet_time = value
                            current_state = 'found_meet'
                        elif value and current_state == 'found_meet':
                            store_run.start_time = value
                            store_run.date = header_value
                            current_state = 'found_start'
                        # Check for anyone scheduled in the office
                        elif value and current_state == 'searching' and 'office' in value.lower() and 'leave' not in value.lower():
                            store_run = StoreRun(
                                date=None, start_time=None)
                            store_run.date = header_value
                            store_run.add_store_name(value)
                            current_state = 'empty_value'
                        # If the current state is 'searching' or 'found_meet', capture the start time
                        elif value and current_state == 'searching':
                            store_run = StoreRun(
                                date=None, start_time=None)
                            store_run.start_time = value
                            store_run.date = header_value
                            current_state = 'found_start'
                        # If the current state is 'found_start', capture the inventory type
                        elif current_state == 'found_start':
                            store_run.add_inv_type(value)
                            current_state = 'found_inv_type'
                        # If the current state is 'found_inv_type', capture the store name
                        elif current_state == 'found_inv_type':
                            store_run.add_store_name(value)
                            current_state = 'found_store_name'
                        # If the current state is 'found_store_name', capture the address
                        elif current_state == 'found_store_name':
                            store_run.add_address(value)
                            current_state = 'found_store_address'
                        # If the current state is 'found_store_address', capture the link
                        elif current_state == 'found_store_address':
                            store_run.add_link(value)
                            current_state = 'found_store_link'
                        # If the current state is 'found_store_link', handle 'TO FOLLOW' or note values
                        elif current_state == 'found_store_link':
                            if value == 'TO FOLLOW':
                                current_state = 'to_follow'
                            elif value and 'APPROX' in value:
                                current_state = 'to_follow'
                            elif value:
                                store_run.add_store_note(value)
                                current_state = 'found_store_note'
                            else:
                                current_state = 'empty_value'
                        # If the current state is 'to_follow', capture another inventory type
                        elif current_state == 'to_follow':
                            store_run.add_inv_type(value)
                            current_state = 'found_inv_type'
                        # If the current state is 'found_store_note', set state to 'empty_value'
                        elif current_state == 'found_store_note':
                            if value == 'TO FOLLOW':
                                current_state = 'to_follow'
                            elif value and 'APPROX' in value:
                                current_state = 'to_follow'
                            elif value:
                                store_run.add_store_note(value, concat=True)
                            else:
                                current_state = 'empty_value'
                        # If the current state is 'empty_value', capture employee details
                        elif current_state == 'empty_value':
                            if 'Office' in store_run.store_name:
                                store_run.add_employee(
                                    value, number_value, note_value, 'none')
                                current_state = 'searching'
                                next_cell = sheet.cell(
                                    row=cell.row + 1, column=cell.column).value
                                if next_cell:
                                    store_run.add_employee(
                                        next_cell, number_value, note_value, 'none')
                                    save_store_runs_to_json()
                                    current_state = 'done'
                                else:
                                    save_store_runs_to_json()
                                    current_state = 'done'
                            else:
                                current_state = 'found_employee'
                                store_run.add_employee(
                                    value, number_value, note_value, 'none')
                        # If the current state is 'found_employee', continue capturing employees
                        elif current_state == "found_employee":
                            if value:
                                store_run.add_employee(
                                    value, number_value, note_value, 'none')
                            elif next_store:
                                current_state = 'searching'
                                save_store_runs_to_json()
                            elif next_store is None or next_store == '':
                                max_rows_to_check = 10
                                for i in range(1, max_rows_to_check + 1):
                                    next_store = sheet.cell(
                                        row=cell.row + i, column=cell.column).value
                                    if next_store:
                                        current_state = 'searching'
                                        # save_store_runs_to_json()
                                        break
                                current_state = 'searching'
                                save_store_runs_to_json()

                            else:
                                save_store_runs_to_json()
                                break_outer_loop = True


else:
    print(f"No folder found with the name {folder_name}", "red")


print("\n\n\n")


# Load the JSON data from the files
with open('json/formatted_users.json', 'r') as f:
    users_data = json.load(f)

with open('json/store_runs.json', 'r') as f:
    store_runs_data = json.load(f)

# Create a mapping of employee display names (in lowercase) to their office locations
employee_office_mapping = {user['displayName'].lower(
): user['office'] for user in users_data if 'displayName' in user and 'office' in user}

# Replace 'none' values in the store_runs employee_list with corresponding office values from the mapping
for run in store_runs_data:
    if 'employee_list' in run:
        for employee, details in run['employee_list'].items():
            if employee.lower() in employee_office_mapping:
                details[-1] = employee_office_mapping[employee.lower()]

# Save the updated store_runs data back to a new JSON file
updated_store_runs_path = 'json/store_runs.json'
with open(updated_store_runs_path, 'w') as f:
    json.dump(store_runs_data, f, indent=4)


# Check for errors
# Load the JSON data from the file
with open(json_file_path, 'r') as file:
    store_runs_data = json.load(file)


with open('json/formatted_users.json', 'r') as file:
    users_data = json.load(file)

# Extract display names and first names from users data
valid_display_names = {user['displayName'].lower() for user in users_data}
valid_first_names = {user['firstName'].lower() for user in users_data}

# Define the validation functions


def validate_meet_times(store_run):
    valid_meet_times = ["meet", "leave"]
    meet_times = store_run.get('meet_time', [])
    if isinstance(meet_times, str):
        meet_times = [meet_times]
    for meet_time in meet_times:
        if not any(vit in meet_time.lower() for vit in valid_meet_times):
            return f"Invalid meet time: {meet_time}"
    return None


def validate_start_time(store_run):
    pattern = re.compile(r'\b\d{1,2}:\d{2}\b')
    start_time = store_run.get('start_time', "")
    if start_time and not pattern.search(start_time):
        return f"Invalid start time: {start_time}"
    return None


def validate_inv_types(store_run):
    valid_inv_types = ["modas", "excel", "dc5"]
    for inv_type in store_run.get('inv_type', []):
        if not any(vit in inv_type.lower() for vit in valid_inv_types):
            return f"Invalid inventory type: {inv_type}"
    return None


def validate_address(address):
    # This regex is designed to match a wide range of address formats
    address_pattern = re.compile(
        r'^\d+\s[A-Za-z0-9\s,.#-]+$|^[A-Za-z]+\d+\s[A-Za-z0-9\s,.#-]+$'
    )
    for address in store_run.get('address', []):
        if address and not address_pattern.match(address):
            return f"Invalid address: {address}"
    return None


def validate_links(store_run):
    for link in store_run.get('link', []):
        if "https" not in link:
            return f"Invalid link: {link}"
    return None


def validate_store_note(store_run):
    employees = store_run.get('employee_list', [])
    for store_note in store_run.get('store_note', []):
        if "https" in store_note:
            return f"Error - link found in store note: {store_note}"
        for employee in employees:
            employee_lower = employee.lower()
            if employee_lower in store_note.lower():
                return f"Warning - employee name found in store note: {employee}"
    return None


def validate_employee_list(store_run):
    employees = store_run.get('employee_list', [])
    if not employees and any(store_run.get(key) for key in ['meet_time', 'start_time', 'inv_type', 'link']):
        return "Employee list is empty when other data is present."
    for employee in employees:
        employee_lower = employee.lower()
        if not any(employee_lower in name for name in valid_display_names) and employee_lower not in valid_first_names:
            return f"Invalid employee: {employee}"
    return None


# Validate the store runs and print errors
for store_run in store_runs_data:
    errors = []

    error = validate_meet_times(store_run)
    if error:
        errors.append(error)

    error = validate_start_time(store_run)
    if error:
        errors.append(error)

    error = validate_inv_types(store_run)
    if error:
        errors.append(error)

    error = validate_links(store_run)
    if error:
        errors.append(error)

    error = validate_address(store_run)
    if error:
        errors.append(error)

    error = validate_store_note(store_run)
    if error:
        errors.append(error)

    error = validate_employee_list(store_run)
    if error:
        errors.append(error)

    if errors:
        print("-------------------------------")
        print(
            f"Errors in store run: {store_run['date']}:")
        for error in errors:
            if 'warning' in error.lower():
                print_colored(f"  - {error}", "yellow")
            else:
                print_colored(f"  - {error}", "red")

print(
    "-------------------------------\n\n\nValidation complete\nUpdating website...\n\n\n")
